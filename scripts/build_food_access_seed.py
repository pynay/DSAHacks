"""One-off builder: extract La Jolla census-tract rows from the real USDA Food Access
Research Atlas (FARA) files into seeds/food_access_la_jolla.csv.

Real data only: values come from the real FARA workbooks and the Census 2010 ZCTA->tract
crosswalk. Where a vintage genuinely lacks a field (e.g. 2010 has no TractSNAP) or marks a
tract 'NULL' (2019/LRAM), the seed cell is left blank and the reason is recorded in `note`
- never faked as 0.

Reads the workbooks with openpyxl in read_only streaming mode so it stays within a few
hundred MB of RAM despite the 147-column, ~72k-row sheets.

Inputs (auto-downloaded to raw/fara/ if absent):
  - raw/fara/FoodAccessResearchAtlasData2010.xlsx   (from the 2010 FARA zip)
  - raw/fara/FoodAccessResearchAtlasData2015.xlsx   (from the 2015 FARA zip)
  - raw/fara/FoodAccessResearchAtlasData2019.xlsx   (the 2019 LRAM xlsx, on 2010 tracts)
  - raw/fara/zcta_tract_rel_10.txt                  (Census 2010 ZCTA<->tract relationship)

Run from repo root:  python scripts/build_food_access_seed.py
"""
import csv
import io
import sys
import zipfile
from pathlib import Path

import openpyxl
import requests

ROOT = Path(__file__).resolve().parent.parent
FARA_DIR = ROOT / "raw" / "fara"
OUT = ROOT / "seeds" / "food_access_la_jolla.csv"
FARA_PAGE = "https://www.ers.usda.gov/data-products/food-access-research-atlas/download-the-data"
ERS = "https://www.ers.usda.gov"
CENSUS_REL = "https://www2.census.gov/geo/docs/maps-data/data/rel/zcta_tract_rel_10.txt"
LA_JOLLA_ZCTA = "92037"
SHEET = "Food Access Research Atlas"

# vintage -> (local xlsx name, remote url, is_zip, member-inside-zip)
SOURCES = {
    2010: ("FoodAccessResearchAtlasData2010.xlsx",
           f"{ERS}/media/5624/2010-food-access-research-atlas-fara-data-and-documentation.zip",
           True, "FoodAccessResearchAtlasData2010.xlsx"),
    2015: ("FoodAccessResearchAtlasData2015.xlsx",
           f"{ERS}/media/5623/2015-food-access-research-atlas-fara-data-and-documentation.zip",
           True, "FoodAccessResearchAtlasData2015.xlsx"),
    2019: ("FoodAccessResearchAtlasData2019.xlsx",
           f"{ERS}/media/5626/2019-large-retailer-access-map-lram-formerly-known-as-the-food-access-research-atlas-fara-data.xlsx",
           False, None),
}

TRACT_CANDS = ["CensusTract", "CensusTractNumber", "GEOID", "GEOID10"]
# Raw metrics read verbatim from FARA. `low_access_pop_share` is intentionally NOT read
# here: FARA's own lapop1share column is a fraction (0-1) in the 2010/2015 vintages but a
# percentage (0-100) in 2019/LRAM, so it is not comparable across vintages. We derive the
# share ourselves as 100 * low_access_pop / pop_total (percentage, consistent everywhere).
METRIC_CANDS = {
    "pop_total": ["POP2010", "Pop2010", "POP2010E"],
    "low_access_pop": ["lapop1", "LAPOP1_10", "lapop1_10", "LATractpop1"],
    "low_income_low_access_pop": ["lalowi1", "LALOWI1_10", "lalowi1_10"],
    "lila_flag": ["LILATracts_1And10", "LILATracts_1and10", "LILA1_10"],
    "snap_housing_units": ["TractSNAP", "TractSNAP1"],
}
# Seed column order (derived low_access_pop_share sits between count and low-income pop).
METRICS = ["pop_total", "low_access_pop", "low_access_pop_share",
           "low_income_low_access_pop", "lila_flag", "snap_housing_units"]


def _derived_share(pop, lap):
    """low_access_pop_share as a percentage (0-100), or '' if either input is missing."""
    if pop in ("", None) or lap in ("", None):
        return ""
    pop_f = float(pop)
    return "" if pop_f == 0 else round(100.0 * float(lap) / pop_f, 6)


def _clean(v):
    """Real value or '' for missing. FARA marks missing as None or the literal 'NULL'."""
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s == "" or s.upper() == "NULL" else s


def ensure_fara_file(year):
    local_name, url, is_zip, member = SOURCES[year]
    path = FARA_DIR / local_name
    if path.exists():
        return path
    FARA_DIR.mkdir(parents=True, exist_ok=True)
    print(f"  downloading FARA {year} ...")
    r = requests.get(url, timeout=600)
    r.raise_for_status()
    if is_zip:
        with zipfile.ZipFile(io.BytesIO(r.content)) as z:
            path.write_bytes(z.read(member))
    else:
        path.write_bytes(r.content)
    return path


def la_jolla_tracts():
    local = FARA_DIR / "zcta_tract_rel_10.txt"
    if local.exists():
        text = local.read_text()
    else:
        print(f"  downloading Census ZCTA<->tract crosswalk ...")
        r = requests.get(CENSUS_REL, timeout=300)
        r.raise_for_status()
        text = r.text
    rows = csv.DictReader(io.StringIO(text))
    tracts = sorted({row["GEOID"].zfill(11) for row in rows if row.get("ZCTA5") == LA_JOLLA_ZCTA})
    if not tracts:
        sys.exit(f"No tracts found for ZCTA {LA_JOLLA_ZCTA}")
    print(f"  {len(tracts)} La Jolla tracts")
    return tracts


def _first(header, cands):
    return next((c for c in cands if c in header), None)


def read_vintage(year, tracts):
    path = ensure_fara_file(year)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.worksheets[0]
    header = [str(h) if h is not None else "" for h in
              next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    tcol = _first(header, TRACT_CANDS)
    if tcol is None:
        sys.exit(f"{year}: no census-tract column among {TRACT_CANDS}")
    ti = header.index(tcol)
    picked = {m: _first(header, cands) for m, cands in METRIC_CANDS.items()}
    idx = {m: (header.index(c) if c else None) for m, c in picked.items()}
    absent = [m for m, c in picked.items() if c is None]
    ljset = set(tracts)
    rows, matched = [], set()
    for rec in ws.iter_rows(min_row=2, values_only=True):
        v = rec[ti]
        if v is None:
            continue
        g = str(v).split(".")[0].zfill(11)
        if g not in ljset:
            continue
        matched.add(g)
        note = f"FARA {year} ({path.name})"
        if absent:
            note += f"; fields absent this vintage: {','.join(absent)}"
        row = {"vintage_year": year, "census_tract": g, "source_url": FARA_PAGE, "note": note}
        for m, i in idx.items():          # raw FARA metrics
            row[m] = "" if i is None else _clean(rec[i])
        row["low_access_pop_share"] = _derived_share(row["pop_total"], row["low_access_pop"])
        rows.append(row)
    wb.close()
    print(f"  {year}: matched {len(matched)}/{len(tracts)} tracts"
          + (f"; absent fields: {','.join(absent)}" if absent else ""))
    return rows


def main():
    print("Resolving La Jolla tracts:")
    tracts = la_jolla_tracts()
    all_rows = []
    for year in sorted(SOURCES):
        print(f"Reading FARA {year}:")
        all_rows += read_vintage(year, tracts)
    if not all_rows:
        sys.exit("No FARA rows extracted.")
    cols = ["vintage_year", "census_tract", "pop_total", "low_access_pop",
            "low_access_pop_share", "low_income_low_access_pop", "lila_flag",
            "snap_housing_units", "source_url", "note"]
    all_rows.sort(key=lambda r: (r["vintage_year"], r["census_tract"]))
    OUT.parent.mkdir(exist_ok=True)
    with open(OUT, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        w.writerows(all_rows)
    vintages = sorted({r["vintage_year"] for r in all_rows})
    tracts_n = len({r["census_tract"] for r in all_rows})
    print(f"\nWrote {len(all_rows)} rows ({tracts_n} tracts x vintages {vintages}) -> {OUT}")


if __name__ == "__main__":
    main()
